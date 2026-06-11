#!/usr/bin/env python3
"""Convert the practitioner policy QA workbook into JSONL evaluation cases."""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUTPUT = ROOT / "eval" / "policy_xlsx_qa.jsonl"

CLAUSE_RE = re.compile(
    r"(?:제\s*)?\d+(?:-\d+)?조(?:\s*제\s*\d+\s*항)?|별표\s*\d+|<?표\s*\d+>?|유의사항|특별약관|보통약관|특약"
)
NUMBER_RE = re.compile(
    r"\d[\d,]*(?:\.\d+)?\s*(?:개월|년|일|회|%|만원|원|배|일째|영업일|시간)"
)
IMPORTANT_TERMS = [
    "계약 전 알릴의무",
    "계약 후 알릴의무",
    "고지의무",
    "보험금",
    "보험료",
    "계약 해지",
    "해지",
    "청약철회",
    "철회",
    "무효",
    "부활",
    "납입최고",
    "독촉",
    "해약환급금",
    "소멸시효",
    "보상하지",
    "보상",
    "면책",
    "자기부담금",
    "공제금액",
    "입원",
    "통원",
    "비급여",
    "실손의료비",
    "3대비급여",
    "도수치료",
    "체외충격파",
    "증식치료",
    "주사료",
    "자기공명영상",
    "다수보험",
    "비례분담",
    "갱신",
    "재가입",
    "지정대리청구",
    "지정대리청구인",
    "이륜자동차",
    "중지제도",
    "분쟁",
    "금융감독원",
    "금융분쟁조정",
    "보험나이",
    "약관",
    "설명의무",
    "청구권",
]
GENERIC_CLAUSE_TERMS = {"보통약관", "특약", "특별약관"}


def _clean(value: Any) -> str:
    return " ".join(str(value or "").split())


def _unique(values: list[str], *, limit: int | None = None) -> list[str]:
    result: list[str] = []
    for value in values:
        text = _clean(value)
        if text and text not in result:
            result.append(text)
        if limit is not None and len(result) >= limit:
            break
    return result


def _clause_terms(reference: str) -> list[str]:
    terms = []
    for match in CLAUSE_RE.finditer(reference):
        term = match.group(0).replace(" ", "")
        article = re.match(r"((?:제)?\d+(?:-\d+)?조)", term)
        if article:
            term = article.group(1)
        terms.append(term)
    terms = _unique(terms)
    specific_terms = [term for term in terms if term not in GENERIC_CLAUSE_TERMS]
    return specific_terms or terms


def _number_terms(text: str) -> list[str]:
    return _unique([_clean(match.group(0)) for match in NUMBER_RE.finditer(text)], limit=6)


def _important_terms(question: str, answer: str, category: str) -> list[str]:
    del category
    haystack = f"{question} {answer}"
    return _unique([term for term in IMPORTANT_TERMS if term in haystack], limit=8)


def _required_groups(answer: str) -> list[list[str]]:
    groups: list[list[str]] = []
    if "보상하지" in answer or "면책" in answer:
        groups.append(["보상하지", "보상 불가", "지급하지", "면책"])
    if "해지" in answer:
        groups.append(["해지", "계약 해지"])
    if "비례" in answer or "분담" in answer:
        groups.append(["비례", "분담", "비례분담"])
    if "갱신" in answer:
        groups.append(["갱신"])
    if "재가입" in answer:
        groups.append(["재가입"])
    return groups


def convert_workbook(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = []
    for row in worksheet.iter_rows(min_row=4, values_only=True):
        number, category, question, reference, expected, *_ = row
        if number is None or not _clean(question):
            continue
        expected_answer = _clean(expected)
        reference_clause = _clean(reference)
        case = {
            "id": f"policy_xlsx_{int(number):03d}",
            "category": _clean(category),
            "question": _clean(question),
            "doc_sources": ["약관"],
            "reference_clause": reference_clause,
            "expected_answer": expected_answer,
            "required_terms": _important_terms(_clean(question), expected_answer, _clean(category)),
            "required_clause_terms": _clause_terms(reference_clause),
            "required_numbers": _number_terms(expected_answer),
            "required_groups": _required_groups(expected_answer),
            "expected_sources": [],
            "notes": "Converted from practitioner policy QA workbook. Expected answer remains available for manual review.",
        }
        rows.append(case)
    return rows


def write_jsonl(rows: list[dict[str, Any]], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        for row in rows:
            f.write(json.dumps(row, ensure_ascii=False, sort_keys=True))
            f.write("\n")


def main() -> int:
    parser = argparse.ArgumentParser(description="Convert policy QA XLSX to JSONL evaluation cases.")
    parser.add_argument("xlsx", type=Path)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()

    rows = convert_workbook(args.xlsx)
    if not rows:
        raise SystemExit("no cases converted")
    write_jsonl(rows, args.output)
    print(json.dumps({"output": str(args.output), "count": len(rows)}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
