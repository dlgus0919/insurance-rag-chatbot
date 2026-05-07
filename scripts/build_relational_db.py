#!/usr/bin/env python3
"""비급여 표준 모델 XLSX를 SQLite 관계형 자산으로 변환한다."""

from __future__ import annotations

import argparse
from dataclasses import dataclass
from datetime import date, datetime
import sqlite3
import sys
import time
import unicodedata
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from src import config

EXPECTED_COLUMNS = [
    "std_cd",
    "std_cd_nm",
    "mid_category_cd",
    "mid_category_cd_nm",
    "hira_care_type_cd",
    "hira_care_type_cd_nm",
    "ins_care_type_cd",
    "ins_care_type_cd_nm",
    "medical_class_cd",
    "medical_class_cd_nm",
    "item_class_level1cd",
    "item_class_level1cd_nm",
    "item_class_level2cd",
    "item_class_level2cd_nm",
    "pay_opn_cd",
    "pay_opn_cd_nm",
    "notes",
    "remarks",
    "source_name",
    "apply_start_date",
    "apply_end_date",
    "source_date",
    "update_type",
]


@dataclass(frozen=True)
class BuildStats:
    source_rows: int
    inserted_rows: int
    duplicate_std_cd_rows: int
    skipped_blank_std_cd_rows: int
    elapsed_sec: float
    db_size_bytes: int


def _normalize_name(value: str) -> str:
    return unicodedata.normalize("NFC", value)


def resolve_default_xlsx() -> Path:
    """설정 경로 또는 루트 파일명 검색으로 D8 XLSX를 찾는다."""

    configured = config.SPREADSHEET_SOURCES[0].path
    if configured.exists():
        return configured

    for path in config.ROOT_DIR.glob("*.xlsx"):
        if "비급여표준모델" in _normalize_name(path.name):
            return path
    raise FileNotFoundError("비급여 표준 모델 XLSX를 프로젝트 루트에서 찾을 수 없습니다.")


def _cell_to_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def _detect_header(rows: Iterable[tuple[object, ...]]) -> tuple[list[str], int]:
    first = tuple(rows)
    if len(first) < 2:
        raise ValueError("XLSX 헤더를 확인하려면 최소 2행이 필요합니다.")

    first_row = [_cell_to_text(value) for value in first[0]]
    second_row = [_cell_to_text(value) for value in first[1]]
    if first_row == EXPECTED_COLUMNS:
        return first_row, 2
    if second_row == EXPECTED_COLUMNS:
        return second_row, 3
    raise ValueError(f"영문 컬럼 키가 명세와 다릅니다: {second_row}")


def _create_schema(connection: sqlite3.Connection) -> None:
    columns_sql = ",\n            ".join(f"{column} TEXT" for column in EXPECTED_COLUMNS)
    connection.executescript(
        f"""
        DROP TABLE IF EXISTS nonpay_standard;
        CREATE TABLE nonpay_standard (
            {columns_sql}
        );
        CREATE UNIQUE INDEX idx_nonpay_standard_std_cd
            ON nonpay_standard(std_cd);
        CREATE INDEX idx_nonpay_standard_mid_category_cd
            ON nonpay_standard(mid_category_cd);
        CREATE INDEX idx_nonpay_standard_medical_class_cd
            ON nonpay_standard(medical_class_cd);
        CREATE INDEX idx_nonpay_standard_apply_start_date
            ON nonpay_standard(apply_start_date);
        """
    )


def build_database(xlsx_path: Path, db_path: Path, batch_size: int = 5000) -> BuildStats:
    """XLSX의 영문 키 행을 기준으로 nonpay_standard 테이블을 생성한다."""

    started = time.perf_counter()
    db_path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = db_path.with_suffix(db_path.suffix + ".tmp")
    if tmp_path.exists():
        tmp_path.unlink()

    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    worksheet = workbook["Result 1"] if "Result 1" in workbook.sheetnames else workbook[workbook.sheetnames[0]]
    header_rows = list(worksheet.iter_rows(min_row=1, max_row=2, values_only=True))
    headers, data_start_row = _detect_header(header_rows)
    std_cd_index = headers.index("std_cd")

    connection = sqlite3.connect(tmp_path)
    connection.execute("PRAGMA journal_mode=OFF")
    connection.execute("PRAGMA synchronous=OFF")
    _create_schema(connection)

    placeholders = ", ".join("?" for _ in EXPECTED_COLUMNS)
    insert_sql = f"INSERT OR IGNORE INTO nonpay_standard ({', '.join(EXPECTED_COLUMNS)}) VALUES ({placeholders})"
    batch: list[tuple[str, ...]] = []
    source_rows = 0
    skipped_blank_std_cd_rows = 0

    for row in worksheet.iter_rows(min_row=data_start_row, values_only=True):
        values = tuple(_cell_to_text(value) for value in row[: len(EXPECTED_COLUMNS)])
        if len(values) < len(EXPECTED_COLUMNS):
            values = values + ("",) * (len(EXPECTED_COLUMNS) - len(values))
        source_rows += 1
        if not values[std_cd_index]:
            skipped_blank_std_cd_rows += 1
            continue
        batch.append(values)
        if len(batch) >= batch_size:
            connection.executemany(insert_sql, batch)
            batch = []

    if batch:
        connection.executemany(insert_sql, batch)

    inserted_rows = connection.execute("SELECT COUNT(*) FROM nonpay_standard").fetchone()[0]
    connection.commit()
    connection.close()
    workbook.close()

    tmp_path.replace(db_path)
    elapsed_sec = time.perf_counter() - started
    duplicate_std_cd_rows = source_rows - skipped_blank_std_cd_rows - inserted_rows
    return BuildStats(
        source_rows=source_rows,
        inserted_rows=inserted_rows,
        duplicate_std_cd_rows=duplicate_std_cd_rows,
        skipped_blank_std_cd_rows=skipped_blank_std_cd_rows,
        elapsed_sec=elapsed_sec,
        db_size_bytes=db_path.stat().st_size,
    )


def main() -> int:
    parser = argparse.ArgumentParser(description="비급여 표준 모델 XLSX를 SQLite로 적재")
    parser.add_argument("--xlsx", type=Path, default=None, help="입력 XLSX 경로")
    parser.add_argument("--output", type=Path, default=config.STANDARD_CODES_DB_PATH, help="출력 SQLite 경로")
    parser.add_argument("--batch-size", type=int, default=5000)
    args = parser.parse_args()

    xlsx_path = args.xlsx or resolve_default_xlsx()
    stats = build_database(xlsx_path=xlsx_path, db_path=args.output, batch_size=args.batch_size)
    print(f"XLSX: {xlsx_path}")
    print(f"SQLite: {args.output}")
    print(f"source_rows: {stats.source_rows:,}")
    print(f"inserted_rows: {stats.inserted_rows:,}")
    print(f"duplicate_std_cd_rows: {stats.duplicate_std_cd_rows:,}")
    print(f"skipped_blank_std_cd_rows: {stats.skipped_blank_std_cd_rows:,}")
    print(f"db_size_mb: {stats.db_size_bytes / 1024 / 1024:.1f}")
    print(f"elapsed_sec: {stats.elapsed_sec:.1f}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
