"""Run a batch claim-calculation evaluation for manually extracted receipts."""

from __future__ import annotations

import argparse
import csv
import json
import subprocess
import sys
from dataclasses import asdict
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from src.claim_calculation.models import ClaimCaseContext, ClaimItemInput
from src.claim_calculation.pipeline import run_claim_calculation

DEFAULT_INPUT = Path(
    "data/hospital_receipts/manual_20260609/manual_extraction/claim_calculation_input.json"
)
DEFAULT_REPORT_ROOT = Path("reports/claim_batch/manual_20260609")
OUTPUT_COLUMNS = [
    "line_id",
    "source_file",
    "page_label",
    "source_row_id",
    "item_group",
    "service_date",
    "input_code",
    "input_name",
    "category",
    "claimed_amount",
    "insured_copay_amount",
    "nonpay_amount",
    "deductible",
    "payable_amount",
    "human_task_amount",
    "calculation_status",
    "excluded_from_calculation",
    "requires_review",
    "review_reasons",
    "rule_summary",
    "ready_for_auto_calculation",
    "practitioner_grade",
    "practitioner_comment",
    "corrected_payable_amount",
]


def _join(value: Any) -> str:
    if isinstance(value, list):
        return ", ".join(str(item) for item in value if item is not None)
    return str(value or "")


def _text(value: Any, default: str = "") -> str:
    if value is None:
        return default
    return str(value)


def _policy_generation(value: Any) -> str:
    text = _text(value, "5th").strip()
    return text if text in {"4th", "5th"} else "5th"


def _zeroish(value: Any) -> bool:
    if value is None:
        return True
    text = str(value).replace(",", "").replace("원", "").strip()
    if not text:
        return True
    try:
        return Decimal(text) == 0
    except (InvalidOperation, ValueError):
        return False


def _split_amounts(raw: dict[str, Any]) -> tuple[str, str]:
    insured = raw.get("insured_copay_amount")
    nonpay = raw.get("nonpay_amount")
    if _zeroish(insured) and _zeroish(nonpay):
        return "", ""
    return _text(insured), _text(nonpay)


def build_claim_items(payload: dict[str, Any]) -> tuple[list[ClaimItemInput], dict[str, dict[str, Any]]]:
    items: list[ClaimItemInput] = []
    metadata: dict[str, dict[str, Any]] = {}
    for index, raw in enumerate(payload.get("claim_items") or [], 1):
        line_id = _text(raw.get("line_id"), f"line-{index:03d}")
        extra_info = raw.get("extra_info") or {}
        insured_copay_amount, nonpay_amount = _split_amounts(raw)
        items.append(
            ClaimItemInput(
                line_id=line_id,
                input_name=_text(raw.get("input_name")),
                input_code=_text(raw.get("input_code")),
                claimed_amount=_text(raw.get("claimed_amount"), "0"),
                insured_copay_amount=insured_copay_amount,
                nonpay_amount=nonpay_amount,
                quantity=_text(raw.get("quantity"), "1"),
                user_category_hint=_text(raw.get("user_category_hint")),
                extra_info=json.dumps(extra_info, ensure_ascii=False),
                is_prescription=bool(raw.get("is_prescription")),
            )
        )
        metadata[line_id] = {
            **extra_info,
            "ready_for_auto_calculation": bool(raw.get("ready_for_auto_calculation")),
        }
    if not items:
        raise ValueError("claim_items is empty")
    return items, metadata


def build_claim_context(
    payload: dict[str, Any],
    *,
    policy_generation: str | None = None,
) -> ClaimCaseContext:
    raw = payload.get("claim_case_context") or {}
    return ClaimCaseContext(
        treatment_date=_text(raw.get("treatment_date")),
        visit_type=_text(raw.get("visit_type")),
        coverage_topic=_text(raw.get("coverage_topic")),
        diagnosis_code=_join(raw.get("diagnosis_code")),
        diagnosis_name=_join(raw.get("diagnosis_name")),
        accident_type=_text(raw.get("accident_type")),
        situation_note=_text(raw.get("situation_note")),
        policy_generation=_policy_generation(policy_generation or raw.get("policy_generation")),
        complication_asserted=bool(raw.get("complication_asserted")),
        same_disease_claimed=bool(raw.get("same_disease_claimed")),
        same_treatment_purpose_claimed=bool(raw.get("same_treatment_purpose_claimed")),
        recurrent_or_continuing_treatment=bool(raw.get("recurrent_or_continuing_treatment")),
        newly_found_disease_claimed=bool(raw.get("newly_found_disease_claimed")),
        treatment_purpose=_text(raw.get("treatment_purpose")),
        evidence_tags=list(raw.get("evidence_tags") or []),
        facility_type=_text(raw.get("facility_type")),
        facility_grade=_text(raw.get("facility_grade")),
    )


def _csv_value(value: Any) -> str:
    if isinstance(value, list):
        return "; ".join(str(item) for item in value)
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    if isinstance(value, bool):
        return "true" if value else "false"
    return "" if value is None else str(value)


def flatten_line_results(
    line_results: list[dict[str, Any]],
    metadata: dict[str, dict[str, Any]],
) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for result in line_results:
        line_id = _text(result.get("line_id"))
        meta = metadata.get(line_id, {})
        row = {column: "" for column in OUTPUT_COLUMNS}
        for column in OUTPUT_COLUMNS:
            if column in result:
                row[column] = _csv_value(result[column])
            elif column in meta:
                row[column] = _csv_value(meta[column])
        row["practitioner_grade"] = ""
        row["practitioner_comment"] = ""
        row["corrected_payable_amount"] = ""
        rows.append(row)
    return rows


def _write_json(path: Path, data: Any) -> None:
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def _write_csv(path: Path, rows: list[dict[str, str]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=OUTPUT_COLUMNS)
        writer.writeheader()
        writer.writerows(rows)


def _write_xlsx(path: Path, rows: list[dict[str, str]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "line_results"
    ws.append(OUTPUT_COLUMNS)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for row in rows:
        ws.append([row.get(column, "") for column in OUTPUT_COLUMNS])
    ws.freeze_panes = "A2"
    widths = {
        "A": 20,
        "B": 34,
        "H": 42,
        "T": 60,
        "W": 20,
        "X": 36,
        "Y": 22,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    wb.save(path)


def _git_version() -> str:
    try:
        result = subprocess.run(
            ["git", "tag", "--points-at", "HEAD"],
            check=False,
            capture_output=True,
            text=True,
        )
    except OSError:
        return "unknown"
    tags = [line.strip() for line in result.stdout.splitlines() if line.strip()]
    return tags[0] if tags else "unknown"


def _default_output_dir(version: str) -> Path:
    stamp = datetime.now().strftime("%Y%m%dT%H%M%S")
    suffix = version if version != "unknown" else "untagged"
    return DEFAULT_REPORT_ROOT / f"{stamp}_{suffix}"


def run_batch(
    *,
    input_path: Path,
    output_dir: Path | None,
    policy_generation: str | None = None,
) -> dict[str, Any]:
    payload = json.loads(input_path.read_text(encoding="utf-8"))
    items, metadata = build_claim_items(payload)
    context = build_claim_context(payload, policy_generation=policy_generation)
    version = _git_version()
    target_dir = output_dir or _default_output_dir(version)
    target_dir.mkdir(parents=True, exist_ok=True)

    result = run_claim_calculation(
        rag_pipeline=None,
        items=items,
        context=context,
        basis_mode="auto",
        use_fake_planner=True,
    )
    rows = flatten_line_results(result.line_results, metadata)
    summary = {
        "app_version": version,
        "input_path": str(input_path),
        "output_dir": str(target_dir),
        "policy_generation": context.policy_generation,
        "input_item_count": len(items),
        "line_result_count": len(result.line_results),
        "ready_for_auto_calculation_count": sum(
            1 for item in metadata.values() if item.get("ready_for_auto_calculation")
        ),
        "review_line_count": sum(1 for row in result.line_results if row.get("requires_review")),
        "excluded_line_count": sum(
            1 for row in result.line_results if row.get("excluded_from_calculation")
        ),
        "claimed_amount": result.claimed_amount,
        "payable_amount": result.payable_amount,
        "deductible": result.deductible,
        "requires_review": result.requires_review,
        "review_reasons": result.review_reasons,
    }

    _write_json(
        target_dir / "input_payload.json",
        {"context": asdict(context), "items": [asdict(item) for item in items]},
    )
    _write_json(target_dir / "claim_response.json", asdict(result))
    _write_csv(target_dir / "line_results.csv", rows)
    _write_xlsx(target_dir / "practitioner_scoring.xlsx", rows)
    _write_json(target_dir / "summary.json", summary)
    (target_dir / "README.md").write_text(
        "\n".join(
            [
                "# Hospital Receipt Claim Batch Evaluation",
                "",
                f"- app_version: `{version}`",
                f"- input_item_count: `{len(items)}`",
                f"- line_result_count: `{len(result.line_results)}`",
                f"- payable_amount: `{result.payable_amount}`",
                f"- deductible: `{result.deductible}`",
                "",
                "`practitioner_scoring.xlsx`의 practitioner_* 컬럼에 실무자 채점 결과를 입력합니다.",
            ]
        ),
        encoding="utf-8",
    )
    return summary


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--output-dir", type=Path)
    parser.add_argument("--policy-generation", choices=["4th", "5th"])
    return parser


def main() -> None:
    args = build_parser().parse_args()
    summary = run_batch(
        input_path=args.input,
        output_dir=args.output_dir,
        policy_generation=args.policy_generation,
    )
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
